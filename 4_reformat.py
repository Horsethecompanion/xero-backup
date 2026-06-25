"""
Step 4: Reformat the raw Xero backup into something human-browsable.

Run this AFTER 3_backup.py has finished (or partially finished -- it's
safe to run on an in-progress backup and re-run later once more data
has downloaded).

What it does:
  1. Renames invoice PDFs and attachments to a readable pattern:
     YYYY-MM-DD_ContactName_Type_OriginalFilename.ext
  2. Builds one flat CSV (transactions.csv) combining Invoices, Bills,
     and Bank Transactions as line-item rows -- easy to open in Excel,
     filter, and pivot.
  3. Builds a simple multi-tab summary workbook (summary.xlsx) with
     Invoices, Bills, Bank Transactions, and Contacts tabs.

Leaves the original xero_backup/ folder untouched -- everything new
goes into a separate xero_backup_organised/ folder alongside it.

Run: python3 4_reformat.py
Requires: pip3 install openpyxl
"""

import json
import re
import shutil
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font
except ImportError:
    openpyxl = None

SRC = Path("xero_backup")
OUT = Path("xero_backup_organised")

DATA_DIR = SRC / "data"
PDF_DIR = SRC / "invoice_pdfs"
ATTACH_DIR = SRC / "attachments"

OUT_FILES = OUT / "files"
OUT_INVOICES = OUT_FILES / "invoices"
OUT_BILLS = OUT_FILES / "bills"
OUT_BANKTXN = OUT_FILES / "bank_transactions"
OUT_CONTACTS = OUT_FILES / "contacts"

for d in (OUT_INVOICES, OUT_BILLS, OUT_BANKTXN, OUT_CONTACTS):
    d.mkdir(parents=True, exist_ok=True)


def load_json(name):
    path = DATA_DIR / f"{name}.json"
    if not path.exists():
        return []
    with open(path) as f:
        return json.load(f)


def xero_date(s):
    """Convert '/Date(1763856000000)/' to a 'YYYY-MM-DD' string. Falls back gracefully."""
    if not s:
        return "0000-00-00"
    match = re.search(r"/Date\((\d+)", s)
    if not match:
        return "0000-00-00"
    millis = int(match.group(1))
    return datetime.fromtimestamp(millis / 1000, tz=timezone.utc).strftime("%Y-%m-%d")


def safe_name(s, max_len=80):
    s = re.sub(r"[^\w\-. ]", "_", str(s))
    s = re.sub(r"\s+", "-", s.strip())
    return s[:max_len]


# ---------------------------------------------------------------------------
# Load core data
# ---------------------------------------------------------------------------

invoices = load_json("Invoices")
contacts = load_json("Contacts")
bank_txns = load_json("BankTransactions")
manual_journals = load_json("ManualJournals")
payments = load_json("Payments")

contacts_by_id = {c["ContactID"]: c for c in contacts if "ContactID" in c}


# ---------------------------------------------------------------------------
# 1. Rename + copy invoice PDFs and attachments into readable folders
# ---------------------------------------------------------------------------

def organise_invoice_pdfs():
    print("Organising invoice/bill PDFs...")
    count = 0
    for inv in invoices:
        pdf_src = PDF_DIR / f"{safe_name(inv.get('InvoiceNumber', inv['InvoiceID']))}.pdf"
        if not pdf_src.exists():
            continue
        date_str = xero_date(inv.get("Date"))
        contact_name = safe_name(inv.get("Contact", {}).get("Name", "Unknown"))
        doc_type = "Invoice" if inv.get("Type") == "ACCREC" else "Bill"
        dest_dir = OUT_INVOICES if doc_type == "Invoice" else OUT_BILLS
        dest_name = f"{date_str}_{contact_name}_{doc_type}_{safe_name(inv.get('InvoiceNumber', ''))}.pdf"
        shutil.copy2(pdf_src, dest_dir / dest_name)
        count += 1
    print(f"  -> {count} PDFs organised")


def organise_attachments():
    print("Organising attachments (receipts/scans)...")
    count = 0

    # Invoice/bill attachments
    inv_attach_dir = ATTACH_DIR / "Invoices"
    if inv_attach_dir.exists():
        inv_by_number = {safe_name(i.get("InvoiceNumber", i["InvoiceID"])): i for i in invoices}
        for f in inv_attach_dir.iterdir():
            if not f.is_file():
                continue
            # files are named "{InvoiceNumber}__{original filename}"
            parts = f.name.split("__", 1)
            inv_number = parts[0]
            orig_name = parts[1] if len(parts) > 1 else f.name
            inv = inv_by_number.get(inv_number)
            if inv:
                date_str = xero_date(inv.get("Date"))
                contact_name = safe_name(inv.get("Contact", {}).get("Name", "Unknown"))
                doc_type = "Invoice" if inv.get("Type") == "ACCREC" else "Bill"
                dest_dir = OUT_INVOICES if doc_type == "Invoice" else OUT_BILLS
                dest_name = f"{date_str}_{contact_name}_{doc_type}_{inv_number}_ATTACHMENT_{safe_name(orig_name)}"
            else:
                dest_dir = OUT_INVOICES
                dest_name = f"UNKNOWN_{safe_name(f.name)}"
            shutil.copy2(f, dest_dir / dest_name)
            count += 1

    # Bank transaction attachments (these are usually the scanned receipts)
    bt_attach_dir = ATTACH_DIR / "BankTransactions"
    if bt_attach_dir.exists():
        bt_by_id = {bt["BankTransactionID"]: bt for bt in bank_txns}
        for f in bt_attach_dir.iterdir():
            if not f.is_file():
                continue
            parts = f.name.split("__", 1)
            bt_id = parts[0]
            orig_name = parts[1] if len(parts) > 1 else f.name
            bt = bt_by_id.get(bt_id)
            if bt:
                date_str = xero_date(bt.get("Date"))
                contact_name = safe_name(bt.get("Contact", {}).get("Name", "Unknown"))
                txn_type = bt.get("Type", "TXN")
                dest_name = f"{date_str}_{contact_name}_{txn_type}_RECEIPT_{safe_name(orig_name)}"
            else:
                dest_name = f"UNKNOWN_{safe_name(f.name)}"
            shutil.copy2(f, OUT_BANKTXN / dest_name)
            count += 1

    # Contact attachments
    contact_attach_dir = ATTACH_DIR / "Contacts"
    if contact_attach_dir.exists():
        for f in contact_attach_dir.iterdir():
            if not f.is_file():
                continue
            shutil.copy2(f, OUT_CONTACTS / safe_name(f.name))
            count += 1

    print(f"  -> {count} attachments organised")


# ---------------------------------------------------------------------------
# 2. Flat combined CSV (one row per line item, like Numerint's export)
# ---------------------------------------------------------------------------

def build_flat_csv():
    print("Building combined transactions.csv...")
    import csv

    rows = []
    fieldnames = [
        "RecordType", "Date", "DueDate", "Status", "ContactName",
        "DocumentNumber", "Reference", "LineDescription", "AccountCode",
        "Quantity", "UnitAmount", "LineAmount", "TaxAmount",
        "SubTotal", "TotalTax", "Total", "AmountDue", "AmountPaid",
        "CurrencyCode", "HasAttachments", "RecordID",
    ]

    for inv in invoices:
        doc_type = "Invoice" if inv.get("Type") == "ACCREC" else "Bill"
        base = {
            "RecordType": doc_type,
            "Date": xero_date(inv.get("Date")),
            "DueDate": xero_date(inv.get("DueDate")),
            "Status": inv.get("Status"),
            "ContactName": inv.get("Contact", {}).get("Name"),
            "DocumentNumber": inv.get("InvoiceNumber"),
            "Reference": inv.get("Reference"),
            "SubTotal": inv.get("SubTotal"),
            "TotalTax": inv.get("TotalTax"),
            "Total": inv.get("Total"),
            "AmountDue": inv.get("AmountDue"),
            "AmountPaid": inv.get("AmountPaid"),
            "CurrencyCode": inv.get("CurrencyCode"),
            "HasAttachments": inv.get("HasAttachments"),
            "RecordID": inv.get("InvoiceID"),
        }
        line_items = inv.get("LineItems") or [{}]
        for li in line_items:
            row = dict(base)
            row["LineDescription"] = li.get("Description")
            row["AccountCode"] = li.get("AccountCode")
            row["Quantity"] = li.get("Quantity")
            row["UnitAmount"] = li.get("UnitAmount")
            row["LineAmount"] = li.get("LineAmount")
            row["TaxAmount"] = li.get("TaxAmount")
            rows.append(row)

    for bt in bank_txns:
        base = {
            "RecordType": f"BankTxn-{bt.get('Type', '')}",
            "Date": xero_date(bt.get("Date")),
            "DueDate": "",
            "Status": bt.get("Status"),
            "ContactName": bt.get("Contact", {}).get("Name"),
            "DocumentNumber": "",
            "Reference": bt.get("Reference"),
            "SubTotal": bt.get("SubTotal"),
            "TotalTax": bt.get("TotalTax"),
            "Total": bt.get("Total"),
            "AmountDue": "",
            "AmountPaid": "",
            "CurrencyCode": bt.get("CurrencyCode", ""),
            "HasAttachments": bt.get("HasAttachments"),
            "RecordID": bt.get("BankTransactionID"),
        }
        line_items = bt.get("LineItems") or [{}]
        for li in line_items:
            row = dict(base)
            row["LineDescription"] = li.get("Description")
            row["AccountCode"] = li.get("AccountCode")
            row["Quantity"] = li.get("Quantity")
            row["UnitAmount"] = li.get("UnitAmount")
            row["LineAmount"] = li.get("LineAmount")
            row["TaxAmount"] = li.get("TaxAmount")
            rows.append(row)

    for mj in manual_journals:
        base = {
            "RecordType": "ManualJournal",
            "Date": xero_date(mj.get("Date")),
            "DueDate": "",
            "Status": mj.get("Status"),
            "ContactName": "",
            "DocumentNumber": "",
            "Reference": mj.get("Narration"),
            "SubTotal": "",
            "TotalTax": "",
            "Total": "",
            "AmountDue": "",
            "AmountPaid": "",
            "CurrencyCode": "",
            "HasAttachments": mj.get("HasAttachments", False),
            "RecordID": mj.get("ManualJournalID"),
        }
        journal_lines = mj.get("JournalLines") or [{}]
        for jl in journal_lines:
            row = dict(base)
            row["LineDescription"] = jl.get("Description")
            row["AccountCode"] = jl.get("AccountCode")
            row["Quantity"] = ""
            row["UnitAmount"] = ""
            row["LineAmount"] = jl.get("LineAmount")
            row["TaxAmount"] = ""
            rows.append(row)

    out_path = OUT / "transactions.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)

    print(f"  -> {len(rows)} rows written to {out_path}")


# ---------------------------------------------------------------------------
# 3. Summary Excel workbook
# ---------------------------------------------------------------------------

def build_summary_xlsx():
    if openpyxl is None:
        print("Skipping summary.xlsx -- openpyxl not installed.")
        print("Install it with: pip3 install openpyxl")
        return

    print("Building summary.xlsx...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def add_sheet(name, headers, row_iter):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in row_iter:
            ws.append(row)
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    # Invoices tab
    inv_rows = (
        [
            i.get("InvoiceNumber"), xero_date(i.get("Date")), xero_date(i.get("DueDate")),
            i.get("Contact", {}).get("Name"), i.get("Status"), i.get("Total"),
            i.get("AmountDue"), i.get("CurrencyCode"),
        ]
        for i in invoices if i.get("Type") == "ACCREC"
    )
    add_sheet("Invoices", ["Number", "Date", "DueDate", "Contact", "Status", "Total", "AmountDue", "Currency"], inv_rows)

    # Bills tab
    bill_rows = (
        [
            i.get("InvoiceNumber"), xero_date(i.get("Date")), xero_date(i.get("DueDate")),
            i.get("Contact", {}).get("Name"), i.get("Status"), i.get("Total"),
            i.get("AmountDue"), i.get("CurrencyCode"),
        ]
        for i in invoices if i.get("Type") == "ACCPAY"
    )
    add_sheet("Bills", ["Number", "Date", "DueDate", "Contact", "Status", "Total", "AmountDue", "Currency"], bill_rows)

    # Bank transactions tab
    bt_rows = (
        [
            xero_date(bt.get("Date")), bt.get("Type"), bt.get("Contact", {}).get("Name"),
            bt.get("Status"), bt.get("Total"), bt.get("Reference"),
        ]
        for bt in bank_txns
    )
    add_sheet("Bank Transactions", ["Date", "Type", "Contact", "Status", "Total", "Reference"], bt_rows)

    # Contacts tab
    contact_rows = (
        [
            c.get("Name"), c.get("EmailAddress"), c.get("IsCustomer"), c.get("IsSupplier"),
        ]
        for c in contacts
    )
    add_sheet("Contacts", ["Name", "Email", "Is Customer", "Is Supplier"], contact_rows)

    out_path = OUT / "summary.xlsx"
    wb.save(out_path)
    print(f"  -> saved {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not SRC.exists():
        print(f"ERROR: {SRC} not found. Run this from the same folder as 3_backup.py,")
        print("after it has downloaded at least some data.")
        return

    organise_invoice_pdfs()
    organise_attachments()
    build_flat_csv()
    build_summary_xlsx()

    print("\n" + "=" * 60)
    print("REORGANISATION COMPLETE")
    print(f"Output folder: {OUT.resolve()}")
    print("=" * 60)


if __name__ == "__main__":
    main()
